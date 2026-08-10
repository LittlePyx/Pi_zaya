from __future__ import annotations

import csv
import copy
import hashlib
import io
import math
import re
import time
from pathlib import Path
from typing import Any

from kb.research_brief import research_brief_context
from kb.retriever import BM25Retriever
from kb.store import load_all_chunks
from kb.tokenize import tokenize


MATRIX_CELL_FIELDS = (
    "method",
    "dataset_or_experiment",
    "metric",
    "key_result",
    "limitation",
)
COMPARISON_DIMENSIONS = ("task", "dataset", "evaluation_protocol", "metric")
COMPARISON_MODES = ("ranking", "replication")
_MAX_MATRIX_SOURCES = 8
_MAX_CELL_VALUE = 520
_MAX_EVIDENCE_QUOTE = 1_200
_FIELD_SPECS: dict[str, tuple[str, re.Pattern[str]]] = {
    "method": (
        "method methodology approach architecture algorithm model framework pipeline implementation",
        re.compile(
            r"\b(?:method|methodology|approach|architecture|algorithm|model|framework|pipeline|implementation|network|techniques?)\b"
            r"|(?:方法|算法|模型|架构|框架|流程|网络|实现)",
            flags=re.IGNORECASE,
        ),
    ),
    "dataset_or_experiment": (
        "dataset benchmark experimental setup evaluation protocol training test samples simulation hardware",
        re.compile(
            r"\b(?:dataset|benchmark|experiment|experimental|evaluation|protocol|training set|test set|simulation|hardware|samples?)\b"
            r"|(?:数据集|基准|实验|评估|训练集|测试集|仿真|硬件|样本)",
            flags=re.IGNORECASE,
        ),
    ),
    "metric": (
        "metric quantitative evaluation accuracy PSNR SSIM LPIPS RMSE F1 AUC IoU latency FPS runtime",
        re.compile(
            r"\b(?:metric|accuracy|precision|recall|PSNR|SSIM|LPIPS|RMSE|NMSE|MSE|MAE|SNR|FID|F1|AUC|IoU|Dice|FPS|latency|runtime|FLOPs|parameters?|mean average precision)\b"
            r"|(?:指标|准确率|精度|召回率|延迟|运行时间|参数量)",
            flags=re.IGNORECASE,
        ),
    ),
    "key_result": (
        "result results performance achieves improves outperforms quantitative comparison ablation",
        re.compile(
            r"\b(?:result|results|performance|achiev(?:e|es|ed)|improv(?:e|es|ed)|outperform(?:s|ed)?|comparison|ablation)\b"
            r"|(?:结果|性能|达到|提升|优于|超过|对比|消融)",
            flags=re.IGNORECASE,
        ),
    ),
    "limitation": (
        "limitation limitations challenge failure weakness drawback future work remains however although",
        re.compile(
            r"\b(?:limit|limitation|limitations|challenge|failure|weakness|drawback|future work|remain|however|although)\b"
            r"|(?:局限|限制|挑战|失败|不足|缺点|未来工作|仍然|然而)",
            flags=re.IGNORECASE,
        ),
    ),
}
_NUMERIC_RE = re.compile(r"(?<![A-Za-z])[-+]?\d+(?:\.\d+)?\s*(?:%|dB|ms|s|fps|Hz|GB|MB)?", re.IGNORECASE)
_STRONG_METRIC_RE = re.compile(
    r"\b(?:PSNR|SSIM|LPIPS|RMSE|NMSE|MSE|MAE|SNR|FID|F1|AUC|IoU|Dice|FPS|FLOPs)\b"
    r"|\bmAP(?:@[.\d:]+)?\b",
)
_COMPARISON_RESULT_RE = re.compile(
    r"^\s*([-+]?(?:\d+(?:\.\d+)?|\.\d+))\s*(%|dB|ms|s|fps|Hz)?\s*$",
    re.IGNORECASE,
)
_METRIC_ALIASES: dict[str, tuple[tuple[str, ...], str]] = {
    "psnr": (("psnr", "peak signal to noise ratio", "peak signal-to-noise ratio"), "higher"),
    "ssim": (("ssim", "structural similarity", "structural similarity index"), "higher"),
    "lpips": (("lpips", "learned perceptual image patch similarity"), "lower"),
    "rmse": (("rmse", "root mean square error", "root-mean-square error"), "lower"),
    "nmse": (("nmse", "normalized mean square error", "normalized mean squared error"), "lower"),
    "mse": (("mse", "mean square error", "mean squared error"), "lower"),
    "mae": (("mae", "mean absolute error"), "lower"),
    "fid": (("fid", "frechet inception distance", "fréchet inception distance"), "lower"),
    "accuracy": (("accuracy", "top-1 accuracy", "top-5 accuracy"), "higher"),
    "precision": (("precision",), "higher"),
    "recall": (("recall",), "higher"),
    "f1": (("f1", "f1 score", "f1-score"), "higher"),
    "auc": (("auc", "area under the curve"), "higher"),
    "iou": (("iou", "intersection over union"), "higher"),
    "dice": (("dice", "dice score"), "higher"),
    "map": (("map", "mean average precision"), "higher"),
    "snr": (("snr", "signal to noise ratio", "signal-to-noise ratio"), "higher"),
    "fps": (("fps", "frames per second"), "higher"),
    "latency": (("latency",), "lower"),
    "runtime": (("runtime", "run time", "execution time"), "lower"),
}
_REFERENCE_HEADING_RE = re.compile(
    r"(?:^|[/ >])(?:references?|bibliography|literature cited|works cited)(?:$|[/ >])",
    re.IGNORECASE,
)
_METHOD_SIGNAL_RE = re.compile(
    r"\b(?:uses?|using|based on|consists? of|combines?|comprises?|architecture|algorithm|framework|pipeline|network|techniques?)\b"
    r"|\b(?:we|this (?:work|paper|study))\s+(?:propose|present|introduce|develop|design|use|employ|combine|implement|adopt|build|construct)\b"
    r"|(?:采用|使用|基于|结合|由.+组成|提出|设计|实现)",
    re.IGNORECASE,
)
_METHOD_MECHANISM_RE = re.compile(
    r"\b(?:based on|consists? of|combines?|comprises?|architecture|algorithm|framework|pipeline|network|techniques?|implementation|"
    r"sampling scheme|transformation|rasterization|reconstruction method|imaging method)\b"
    r"|\b(?:we|this (?:work|paper|study))\s+(?:propose|present|introduce|develop|design|implement)\b"
    r"|(?:基于|结合|由.+组成|架构|算法|框架|网络|技术|实现|采样方案|变换|重建方法|成像方法|提出|设计)",
    re.IGNORECASE,
)
_EXPERIMENT_SIGNAL_RE = re.compile(
    r"\b(?:dataset|benchmark|experimental setup|evaluation protocol|training set|test set|simulation|hardware|videos?|phantoms?|subjects?)\b"
    r"|\b\d[\d, ]*\s+samples?\b"
    r"|\b(?:comparison )?experiments?\s+(?:use|uses|used|were|are|include|includes|compare|compares)\b"
    r"|\bwe\s+(?:conduct|evaluate|test|train)\b"
    r"|(?:数据集|基准|实验设置|评估协议|训练集|测试集|仿真|硬件|样本|场景|测量)",
    re.IGNORECASE,
)
_RESULT_SIGNAL_RE = re.compile(
    r"\b(?:achiev(?:e|es|ed)|improv(?:e|es|ed)|outperform(?:s|ed)?|surpass(?:es|ed)?|exceed(?:s|ed)?|"
    r"reduc(?:e|es|ed|tion)|increas(?:e|es|ed)|higher|lower|better|superior|state-of-the-art|SOTA|"
    r"demonstrat(?:e|es|ed) that|show(?:s|ed) that)\b"
    r"|(?:达到|提升|优于|超过|降低|减少|增加|更高|更低|结果表明|实验表明)",
    re.IGNORECASE,
)
_LIMITATION_SIGNAL_RE = re.compile(
    r"\b(?:limitation|limitations|challenge|failure|weakness|drawback|trade[- ]?off|at the cost|suffer(?:s|ed)?|"
    r"constrain(?:s|ed|t)?|bottleneck|future work|cannot|unable|fails? to)\b"
    r"|\b(?:however|although|but)\b"
    r"|\bremain(?:s|ed)?\s+(?:limited|challenging|unclear|unknown|difficult)\b"
    r"|(?:局限|限制|挑战|失败|不足|缺点|代价|权衡|瓶颈|未来工作|无法|不能|仍不明确)",
    re.IGNORECASE,
)
_NEGATIVE_LIMIT_SIGNAL_RE = re.compile(
    r"\b(?:limit|limitation|challenge|failure|weakness|drawback|trade[- ]?off|cost|suffer|constraint|bottleneck|"
    r"cannot|unable|fails?|not|difficult|slow|lack|degrad|sacrific|incompatib|unclear|unknown|dependent)\w*\b"
    r"|(?:局限|限制|挑战|失败|不足|缺点|代价|权衡|瓶颈|无法|不能|困难|缓慢|缺少|牺牲|不兼容|不明确|依赖)",
    re.IGNORECASE,
)
_PRIOR_METHOD_RE = re.compile(
    r"\b(?:these|such|existing|previous|prior|conventional|traditional|current|other)\b.{0,60}\b(?:methods?|approaches?|algorithms?|models?|networks?|techniques?|systems?)\b",
    re.IGNORECASE,
)
_CURRENT_WORK_RE = re.compile(
    r"\b(?:we|our|this (?:work|paper|study|method|approach|algorithm|model|network|system)|proposed|presented|introduced)\b",
    re.IGNORECASE,
)
_CURRENT_METHOD_PROPOSAL_RE = re.compile(
    r"\b(?:we|this (?:work|paper|study))\s+(?:propose|present|introduce|develop|design|implement)\b"
    r"|\b(?:our|the proposed|this)\s+(?:method|approach|algorithm|model|network|framework)\b",
    re.IGNORECASE,
)
_RESULT_META_RE = re.compile(
    r"\b(?:results?|values?)\s+(?:are|were)\s+(?:computed|shown|presented|listed|reported)\b"
    r"|\bwe compare\b.{0,100}\b(?:shown?|presented|listed|reported)\s+in\s+(?:table|fig(?:ure)?)\b",
    re.IGNORECASE,
)
_METHOD_META_RE = re.compile(
    r"\b(?:describe|discuss|outline)\b.{0,80}\b(?:how|method|algorithm|implementation)\b",
    re.IGNORECASE,
)
_METHOD_NEGATIVE_RE = re.compile(
    r"\b(?:unrealistic|cannot|unable|not feasible|infeasible|struggle|suffer|limitation)\w*\b",
    re.IGNORECASE,
)
_OTHER_METHOD_LIMIT_RE = re.compile(
    r"\b(?:these|such|this type of|current|existing|previous|prior|other)\b.{0,60}\b(?:methods?|approaches?|algorithms?|models?|networks?|techniques?|systems?)\b"
    r"|\[[0-9,;\-\s]+\]\s+(?:shows?|reports?|finds?)\b"
    r"|\bsome works?\b.{0,40}\b(?:claim|show|report|find)s?\b"
    r"|\b[A-Z][A-Za-z0-9-]{2,}\s+(?:struggles?|suffers?|fails?|has limitations?)\b",
    re.IGNORECASE,
)
_LIMITATION_RESOLUTION_RE = re.compile(
    r"\b(?:address|overcome|resolve|alleviate|mitigate)\w*\b.{0,60}\b(?:limit|limitation|challenge|drawback)s?\b"
    r"|\b(?:limit|limitation|challenge|drawback)s?\b.{0,60}\b(?:are|is|were|was)?\s*(?:addressed|overcome|resolved|alleviated|mitigated)\b",
    re.IGNORECASE,
)
_NON_LIMITATION_CHALLENGE_RE = re.compile(
    r"\bchallenge\s+(?:track\d*|dataset|benchmark|competition)\b",
    re.IGNORECASE,
)
_LIMITATION_ABSENCE_RE = re.compile(
    r"\b(?:no|without|not)\s+(?:(?:known|reported|explicit|significant|major|clear|apparent|any|a|an)\s+){0,3}"
    r"(?:limits?|limitations?|drawbacks?|challenges?|weaknesses?)\b"
    r"|\bwithout\s+(?:reporting|mentioning|identifying|showing)\s+(?:any\s+|a\s+|an\s+)?"
    r"(?:limits?|limitations?|drawbacks?|challenges?|weaknesses?)\b",
    re.IGNORECASE,
)
_DATASET_META_RE = re.compile(
    r"\b(?:images? for the dataset|dataset|results?)\b.{0,50}\b(?:shown|presented|listed)\s+in\s+(?:fig(?:ure)?|table)\b"
    r"|\b(?:images? for the )?dataset\b.{0,50}\bsee\b"
    r"|\bsee\s+(?:the\s+)?supplementary\b"
    r"|\bdetails\b.{0,100}\b(?:found|provided|described)\b",
    re.IGNORECASE,
)
_CAPTION_ONLY_RE = re.compile(
    r"^(?:\([a-z]\)\s+|!\[)?(?:\*+)?\s*(?:the\s+)?(?:modified\s+)?(?:figure|fig\.?|table|visualization|experimental setup)\s*\d*\b",
    re.IGNORECASE,
)
_FIELD_HEADING_RE: dict[str, re.Pattern[str]] = {
    "method": re.compile(r"\b(?:method|methodology|approach|architecture|model|framework|implementation|abstract)\b|(?:方法|模型|架构|实现|摘要)", re.I),
    "dataset_or_experiment": re.compile(r"\b(?:experiment|evaluation|dataset|setup|implementation|results?)\b|(?:实验|评估|数据集|设置|实现|结果)", re.I),
    "metric": re.compile(r"\b(?:metric|quantitative|evaluation|experiment|results?)\b|(?:指标|定量|评估|实验|结果)", re.I),
    "key_result": re.compile(r"\b(?:result|evaluation|experiment|comparison|conclusion|abstract)\b|(?:结果|评估|实验|对比|结论|摘要)", re.I),
    "limitation": re.compile(r"\b(?:limitation|discussion|conclusion|future|outlook)\b|(?:局限|讨论|结论|未来|展望)", re.I),
}


def _text(value: object, *, limit: int = 2_000, multiline: bool = False) -> str:
    text = str(value or "").replace("\x00", " ")
    if multiline:
        text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    else:
        text = re.sub(r"\s+", " ", text).strip()
    return text[: max(0, int(limit))]


def _source_variants(value: object) -> set[str]:
    raw = _text(value, limit=1_200).replace("\\", "/")
    if not raw:
        return set()
    variants = {raw.lower()}
    try:
        path = Path(raw)
        variants.update(
            {
                path.name.lower(),
                path.stem.lower(),
                str(path.expanduser().resolve(strict=False)).replace("\\", "/").lower(),
            }
        )
    except Exception:
        pass
    return {item for item in variants if item}


def _meta(hit: dict[str, Any]) -> dict[str, Any]:
    value = hit.get("meta")
    return value if isinstance(value, dict) else {}


def _source_path(hit: dict[str, Any]) -> str:
    return _text(_meta(hit).get("source_path"), limit=1_200)


def _source_chunks(chunks: list[dict[str, Any]], source_path: str) -> list[dict[str, Any]]:
    expected = _source_variants(source_path)
    if not expected:
        return []
    matched: list[dict[str, Any]] = []
    for chunk in chunks:
        if not isinstance(chunk, dict):
            continue
        raw = _source_path(chunk).replace("\\", "/")
        if not raw:
            continue
        variants = {raw.lower()}
        try:
            path = Path(raw)
            variants.update({path.name.lower(), path.stem.lower()})
        except Exception:
            pass
        if expected & variants:
            matched.append(chunk)
    return matched


def _sentence_candidates(value: object) -> list[str]:
    raw = _text(value, limit=8_000, multiline=True)
    if not raw:
        return []
    raw = re.sub(r"<!--.*?-->", " ", raw, flags=re.DOTALL)
    candidates: list[str] = []
    for line in raw.splitlines():
        if line.lstrip().startswith("#"):
            continue
        clean_line = re.sub(r"^\s*(?:#{1,6}|[-*+]|\d+[.)])\s*", "", line).strip()
        if not clean_line:
            continue
        pieces = re.split(r"(?<=[.!?。！？；;])\s+", clean_line)
        for piece in pieces:
            clean = _text(piece, limit=_MAX_CELL_VALUE)
            if len(clean) >= 20:
                candidates.append(clean)
    if not candidates:
        fallback = _text(raw, limit=_MAX_CELL_VALUE)
        if fallback:
            candidates.append(fallback)
    return candidates[:24]


def _candidate_score(
    sentence: str,
    *,
    field: str,
    pattern: re.Pattern[str],
    objective_tokens: set[str],
    heading: str,
) -> float:
    if not pattern.search(sentence) and not (field == "metric" and _STRONG_METRIC_RE.search(sentence)):
        return -1.0
    numeric_surface = re.sub(r"\[[0-9,;\-\s]+\]", " ", sentence)
    if field == "metric" and not (_NUMERIC_RE.search(numeric_surface) or _STRONG_METRIC_RE.search(sentence)):
        return -1.0
    if field == "metric" and sentence.count("·") >= 3:
        return -1.0
    if field == "method" and (
        not _METHOD_SIGNAL_RE.search(sentence)
        or not _METHOD_MECHANISM_RE.search(sentence)
    ):
        return -1.0
    if field == "method" and _PRIOR_METHOD_RE.search(sentence) and not _CURRENT_METHOD_PROPOSAL_RE.search(sentence):
        return -1.0
    if field == "method" and (_METHOD_META_RE.search(sentence) or _METHOD_NEGATIVE_RE.search(sentence)):
        return -1.0
    if field == "dataset_or_experiment" and (
        len(sentence) < 36 or not _EXPERIMENT_SIGNAL_RE.search(sentence)
    ):
        return -1.0
    if field == "dataset_or_experiment" and _DATASET_META_RE.search(sentence):
        return -1.0
    if field == "key_result" and not _RESULT_SIGNAL_RE.search(sentence):
        return -1.0
    if field == "key_result" and _RESULT_META_RE.search(sentence):
        return -1.0
    if field == "limitation" and (
        not _LIMITATION_SIGNAL_RE.search(sentence)
        or not _NEGATIVE_LIMIT_SIGNAL_RE.search(sentence)
    ):
        return -1.0
    if field == "limitation" and _OTHER_METHOD_LIMIT_RE.search(sentence) and not _CURRENT_WORK_RE.search(sentence):
        return -1.0
    if field == "limitation" and _LIMITATION_RESOLUTION_RE.search(sentence):
        return -1.0
    if field == "limitation" and _NON_LIMITATION_CHALLENGE_RE.search(sentence):
        return -1.0
    if field == "limitation" and _LIMITATION_ABSENCE_RE.search(sentence):
        return -1.0
    sentence_tokens = {str(token).lower() for token in tokenize(sentence) if str(token).strip()}
    objective_weight = 0.8 if field == "method" else 0.45
    score = 3.0 + min(3.0, float(len(sentence_tokens & objective_tokens)) * objective_weight)
    if field in {"metric", "key_result"} and _NUMERIC_RE.search(numeric_surface):
        score += 2.0
    if field == "limitation" and re.search(r"\b(?:however|although|but|remain)\b|(?:然而|但是|仍然)", sentence, re.I):
        score += 0.5
    if _FIELD_HEADING_RE[field].search(heading):
        score += 1.0
    score += min(len(sentence), 320) / 1_000.0
    return score


def _best_cell_hit(
    chunks: list[dict[str, Any]],
    *,
    field: str,
    objective: str,
    retriever: BM25Retriever | None = None,
    excluded_quotes: set[str] | None = None,
) -> tuple[dict[str, Any], str] | None:
    if not chunks:
        return None
    query_terms, pattern = _FIELD_SPECS[field]
    query = query_terms
    active_retriever = retriever or BM25Retriever(chunks)
    ranked = active_retriever.search(query, top_k=min(20, max(8, len(chunks))))
    if field == "method" and objective:
        contextual = active_retriever.search(
            f"{objective} {query_terms}",
            top_k=min(20, max(8, len(chunks))),
        )
        seen_ranked = {
            str(item.get("id") or _meta(item).get("chunk_id") or id(item))
            for item in ranked
        }
        for item in contextual:
            identity = str(item.get("id") or _meta(item).get("chunk_id") or id(item))
            if identity not in seen_ranked:
                ranked.append(item)
                seen_ranked.add(identity)
    if not ranked:
        ranked = chunks[:10]
    objective_tokens = {
        str(token).lower()
        for token in tokenize(objective)
        if len(str(token).strip()) >= 2
    }
    best: tuple[float, dict[str, Any], str] | None = None
    for rank, hit in enumerate(ranked[:32]):
        meta = _meta(hit)
        heading = _text(meta.get("heading_path") or meta.get("top_heading"), limit=800)
        if _REFERENCE_HEADING_RE.search(heading):
            continue
        for sentence in _sentence_candidates(hit.get("text")):
            if _CAPTION_ONLY_RE.search(sentence):
                continue
            if _normal(sentence) in set(excluded_quotes or set()):
                continue
            score = _candidate_score(
                sentence,
                field=field,
                pattern=pattern,
                objective_tokens=objective_tokens,
                heading=heading,
            )
            if score < 0:
                continue
            if pattern.search(heading):
                score += 0.6
            score += max(0.0, 0.25 - rank * 0.025)
            if best is None or score > best[0]:
                best = (score, hit, sentence)
    return (best[1], best[2]) if best else None


def _evidence_id(source_path: str, field: str, hit: dict[str, Any], quote: str) -> str:
    meta = _meta(hit)
    seed = "|".join(
        (
            source_path.replace("\\", "/").lower(),
            field,
            _text(hit.get("id") or meta.get("chunk_id"), limit=300),
            _text(meta.get("block_id") or meta.get("anchor_id"), limit=300),
            quote,
        )
    )
    return f"ev_{hashlib.sha1(seed.encode('utf-8', errors='ignore')).hexdigest()[:16]}"


def _empty_cell(field: str) -> dict[str, Any]:
    return {
        "field": field,
        "value": "",
        "support_status": "missing",
        "evidence_ids": [],
        "manual_override": False,
    }


def _normal(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _source_identity(value: object) -> str:
    variants = _source_variants(value)
    return sorted(variants, key=lambda item: (-len(item), item))[0] if variants else ""


def _comparison_flags(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    active_rows = [row for row in rows if str(row.get("source_status") or "active") == "active"]
    flags: list[dict[str, Any]] = []
    for field, code, message in (
        (
            "dataset_or_experiment",
            "experimental_conditions_differ",
            "Experimental conditions differ across sources; compare quantitative results cautiously.",
        ),
        (
            "metric",
            "metrics_differ",
            "Reported metrics differ across sources; values are not directly comparable.",
        ),
    ):
        values = {
            _normal((row.get("cells") or {}).get(field, {}).get("value"))
            for row in active_rows
            if isinstance(row.get("cells"), dict)
            and isinstance((row.get("cells") or {}).get(field), dict)
            and _normal((row.get("cells") or {}).get(field, {}).get("value"))
        }
        if len(values) > 1:
            flags.append({"code": code, "severity": "info", "message": message, "field": field})
    return flags


def evidence_matrix_comparison_flags(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Recompute source-comparability warnings without rebuilding matrix cells."""
    return _comparison_flags(rows)


def _comparison_normal(value: object) -> str:
    text = _text(value, limit=4_000, multiline=True).casefold()
    text = re.sub(r"[`*_{}$|]+", " ", text)
    text = text.replace("↑", " higher ").replace("↓", " lower ")
    text = re.sub(r"[^\w.%+\-/]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def _comparison_contains(text: object, value: object) -> bool:
    haystack = _comparison_normal(text)
    return _comparison_contains_normal(haystack, value)


def _comparison_contains_normal(haystack: str, value: object) -> bool:
    needle = _comparison_normal(value)
    if len(needle) < 2 or not haystack:
        return False
    if re.fullmatch(r"[-+]?(?:\d+(?:\.\d+)?|\.\d+)(?:\s*(?:%|db|ms|s|fps|hz))?", needle, re.I):
        return bool(re.search(rf"(?<![\d.]){re.escape(needle)}(?!\d)", haystack, re.I))
    return needle in haystack


def _metric_contract(value: object) -> tuple[str, str]:
    normalized = _comparison_normal(value)
    for canonical, (aliases, direction) in _METRIC_ALIASES.items():
        if any(_comparison_normal(alias) == normalized for alias in aliases):
            return canonical, direction
    return "", ""


def _comparison_result(value: object) -> tuple[float | None, str]:
    match = _COMPARISON_RESULT_RE.fullmatch(_text(value, limit=80))
    if not match:
        return None, ""
    try:
        number = float(match.group(1))
    except (TypeError, ValueError):
        return None, ""
    unit = str(match.group(2) or "").casefold()
    return (number if math.isfinite(number) else None), unit


def _comparison_hit(
    chunks: list[tuple[dict[str, Any], str]],
    values: list[str],
) -> dict[str, Any] | None:
    required = [value for value in values if _comparison_normal(value)]
    if not required:
        return None
    candidates: list[tuple[tuple[int, int], dict[str, Any]]] = []
    for hit, normalized_text in chunks:
        if not all(_comparison_contains_normal(normalized_text, value) for value in required):
            continue
        meta = _meta(hit)
        structured = str(meta.get("structured_kind") or "") in {"table_metric", "table_row"}
        candidates.append(((0 if structured else 1, len(str(hit.get("text") or ""))), hit))
    return min(candidates, key=lambda item: item[0])[1] if candidates else None


def _comparison_evidence_id(
    audit_id: str,
    side: str,
    hit: dict[str, Any],
) -> str:
    meta = _meta(hit)
    seed = "|".join(
        (
            audit_id,
            side,
            _text(hit.get("id") or meta.get("chunk_id"), limit=300),
            _text(meta.get("block_id") or meta.get("anchor_id"), limit=300),
        )
    )
    return f"cev_{hashlib.sha1(seed.encode('utf-8', errors='ignore')).hexdigest()[:16]}"


def _comparison_evidence(
    *,
    audit_id: str,
    side: str,
    row: dict[str, Any],
    hit: dict[str, Any],
    supports: list[str],
) -> dict[str, Any]:
    meta = _meta(hit)
    return {
        "id": _comparison_evidence_id(audit_id, side, hit),
        "comparison_audit_id": audit_id,
        "side": side,
        "supports": sorted(set(supports)),
        "source_item_key": _text(row.get("source_item_key"), limit=500),
        "source_path": _text(row.get("source_path"), limit=1_200),
        "source_name": _text(row.get("source_name") or row.get("paper"), limit=500),
        "title": _text(row.get("paper") or row.get("source_name"), limit=800),
        "heading_path": _text(meta.get("heading_path") or meta.get("top_heading"), limit=800),
        "location_label": _text(meta.get("location_label") or meta.get("heading_path"), limit=500),
        "page_start": meta.get("page_start") or meta.get("page") or None,
        "page_end": meta.get("page_end") or meta.get("page") or None,
        "block_id": _text(meta.get("block_id"), limit=500),
        "anchor_id": _text(meta.get("anchor_id") or meta.get("anchor"), limit=500),
        "evidence_quote": _text(hit.get("text"), limit=_MAX_EVIDENCE_QUOTE, multiline=True),
        "chunk_id": _text(hit.get("id") or meta.get("chunk_id"), limit=500),
        "structured_kind": _text(meta.get("structured_kind"), limit=80),
        "table_metric_direction": _text(meta.get("table_metric_direction"), limit=20),
    }


def _comparison_dimensions(spec: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for item in list(spec.get("dimensions") or []):
        if not isinstance(item, dict):
            continue
        dimension = _text(item.get("dimension"), limit=80)
        if dimension not in COMPARISON_DIMENSIONS or dimension in result:
            continue
        result[dimension] = {
            "dimension": dimension,
            "left_value": _text(item.get("left_value"), limit=240),
            "right_value": _text(item.get("right_value"), limit=240),
            "mapping_confirmed": bool(item.get("mapping_confirmed")),
        }
    return result


def audit_evidence_comparison(
    *,
    rows: list[dict[str, Any]],
    spec: dict[str, Any],
    db_dir: str | Path,
    corpus_chunks: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Audit an explicit, source-paired quantitative comparison without inferring missing facts."""
    total_started = time.perf_counter()
    mode = _text(spec.get("mode"), limit=40).lower()
    if mode not in COMPARISON_MODES:
        mode = "ranking"
    left_row_id = _text(spec.get("left_row_id"), limit=120)
    right_row_id = _text(spec.get("right_row_id"), limit=120)
    dimensions = _comparison_dimensions(spec)
    clean_spec = {
        "mode": mode,
        "left_row_id": left_row_id,
        "right_row_id": right_row_id,
        "dimensions": [dimensions[key] for key in COMPARISON_DIMENSIONS if key in dimensions],
        "left_target": _text(spec.get("left_target"), limit=240),
        "right_target": _text(spec.get("right_target"), limit=240),
        "target_mapping_confirmed": bool(spec.get("target_mapping_confirmed")),
        "left_result": _text(spec.get("left_result"), limit=80),
        "right_result": _text(spec.get("right_result"), limit=80),
    }
    audit_seed = repr(clean_spec)
    audit_id = f"cmp_{hashlib.sha1(audit_seed.encode('utf-8', errors='ignore')).hexdigest()[:16]}"
    by_id = {
        _text(row.get("id"), limit=120): row
        for row in rows
        if isinstance(row, dict) and _text(row.get("id"), limit=120)
    }
    left_row = by_id.get(left_row_id)
    right_row = by_id.get(right_row_id)
    reasons: list[str] = []
    if left_row is None:
        reasons.append("left_row_not_found")
    if right_row is None:
        reasons.append("right_row_not_found")
    if left_row_id and left_row_id == right_row_id:
        reasons.append("comparison_requires_two_rows")
    for dimension in COMPARISON_DIMENSIONS:
        item = dimensions.get(dimension)
        if not item or not item["left_value"] or not item["right_value"]:
            reasons.append(f"missing_{dimension}")
    for key in ("left_target", "right_target", "left_result", "right_result"):
        if not clean_spec[key]:
            reasons.append(f"missing_{key}")

    load_started = time.perf_counter()
    chunks = (
        [item for item in corpus_chunks if isinstance(item, dict)]
        if corpus_chunks is not None
        else [item for item in load_all_chunks(Path(db_dir)) if isinstance(item, dict)]
        if left_row and right_row
        else []
    )
    load_ms = round((time.perf_counter() - load_started) * 1000, 3)
    matching_started = time.perf_counter()
    source_chunks = {
        "left": _source_chunks(chunks, str((left_row or {}).get("source_path") or "")),
        "right": _source_chunks(chunks, str((right_row or {}).get("source_path") or "")),
    }
    source_indexes = {
        side: [(hit, _comparison_normal(hit.get("text"))) for hit in hits]
        for side, hits in source_chunks.items()
    }
    evidence: list[dict[str, Any]] = []
    evidence_by_id: dict[str, dict[str, Any]] = {}
    evidence_by_side_hit: dict[tuple[str, str], dict[str, Any]] = {}
    bindings: dict[str, dict[str, str]] = {"left": {}, "right": {}}

    def bind(side: str, row: dict[str, Any] | None, label: str, values: list[str]) -> None:
        if row is None:
            return
        hit = _comparison_hit(source_indexes[side], values)
        if hit is None:
            reasons.append(f"{side}_{label}_evidence_missing")
            return
        hit_key = _text(hit.get("id") or _meta(hit).get("chunk_id") or id(hit), limit=500)
        key = (side, hit_key)
        item = evidence_by_side_hit.get(key)
        if item is None:
            item = _comparison_evidence(
                audit_id=audit_id,
                side=side,
                row=row,
                hit=hit,
                supports=[label],
            )
            evidence_by_side_hit[key] = item
            evidence.append(item)
            evidence_by_id[str(item["id"])] = item
        else:
            item["supports"] = sorted({*list(item.get("supports") or []), label})
        bindings[side][label] = str(item["id"])

    left_result_terms = [
        str(dimensions.get("dataset", {}).get("left_value") or ""),
        str(dimensions.get("metric", {}).get("left_value") or ""),
        str(clean_spec["left_target"]),
        str(clean_spec["left_result"]),
    ]
    right_result_terms = [
        str(dimensions.get("dataset", {}).get("right_value") or ""),
        str(dimensions.get("metric", {}).get("right_value") or ""),
        str(clean_spec["right_target"]),
        str(clean_spec["right_result"]),
    ]
    if all(left_result_terms):
        bind("left", left_row, "result", left_result_terms)
    if all(right_result_terms):
        bind("right", right_row, "result", right_result_terms)
    for dimension in COMPARISON_DIMENSIONS:
        item = dimensions.get(dimension)
        if not item:
            continue
        for side, row in (("left", left_row), ("right", right_row)):
            value = str(item[f"{side}_value"])
            result_evidence = evidence_by_id.get(bindings[side].get("result", ""))
            if result_evidence and _comparison_contains(result_evidence.get("evidence_quote"), value):
                result_evidence["supports"] = sorted(
                    {*list(result_evidence.get("supports") or []), dimension}
                )
                bindings[side][dimension] = str(result_evidence["id"])
            else:
                bind(side, row, dimension, [value])
    matching_ms = round((time.perf_counter() - matching_started) * 1000, 3)

    validation_started = time.perf_counter()
    dimension_audits: list[dict[str, Any]] = []
    user_confirmed_mappings: list[str] = []
    for dimension in COMPARISON_DIMENSIONS:
        item = dimensions.get(dimension)
        if not item:
            continue
        left_value = str(item["left_value"])
        right_value = str(item["right_value"])
        if dimension == "metric":
            left_metric, _left_direction = _metric_contract(left_value)
            right_metric, _right_direction = _metric_contract(right_value)
            equivalent = bool(left_metric and left_metric == right_metric)
            match_type = "controlled_alias" if equivalent else "mismatch"
        elif _comparison_normal(left_value) == _comparison_normal(right_value):
            equivalent = True
            match_type = "exact"
        elif bool(item["mapping_confirmed"]):
            equivalent = True
            match_type = "user_confirmed"
            user_confirmed_mappings.append(dimension)
        else:
            equivalent = False
            match_type = "mismatch"
        supported = dimension in bindings["left"] and dimension in bindings["right"]
        if not equivalent:
            reasons.append(f"{dimension}_mismatch")
        dimension_audits.append(
            {
                **item,
                "equivalent": equivalent,
                "match_type": match_type,
                "evidence_supported": supported,
                "left_evidence_id": bindings["left"].get(dimension, ""),
                "right_evidence_id": bindings["right"].get(dimension, ""),
            }
        )

    left_metric, metric_direction = _metric_contract(dimensions.get("metric", {}).get("left_value"))
    right_metric, right_direction = _metric_contract(dimensions.get("metric", {}).get("right_value"))
    if not left_metric or left_metric != right_metric or metric_direction != right_direction:
        reasons.append("unsupported_or_mismatched_metric")
    result_evidence_ids = {bindings[side].get("result", "") for side in ("left", "right")}
    observed_directions = {
        "higher" if "↑" in str(item.get("table_metric_direction") or "") else "lower"
        for item in evidence
        if str(item.get("id") or "") in result_evidence_ids
        and (
            "↑" in str(item.get("table_metric_direction") or "")
            or "↓" in str(item.get("table_metric_direction") or "")
        )
    }
    if observed_directions and (len(observed_directions) > 1 or metric_direction not in observed_directions):
        reasons.append("metric_direction_conflict")

    left_number, left_unit = _comparison_result(clean_spec["left_result"])
    right_number, right_unit = _comparison_result(clean_spec["right_result"])
    if left_number is None:
        reasons.append("left_result_not_numeric")
    if right_number is None:
        reasons.append("right_result_not_numeric")
    if left_unit != right_unit:
        reasons.append("result_unit_mismatch")
    if "result" not in bindings["left"]:
        reasons.append("left_result_not_jointly_supported")
    if "result" not in bindings["right"]:
        reasons.append("right_result_not_jointly_supported")

    target_match_type = "not_required"
    if mode == "replication":
        if _comparison_normal(clean_spec["left_target"]) == _comparison_normal(clean_spec["right_target"]):
            target_match_type = "exact"
        elif clean_spec["target_mapping_confirmed"]:
            target_match_type = "user_confirmed"
            user_confirmed_mappings.append("comparison_target")
        else:
            target_match_type = "mismatch"
            reasons.append("comparison_target_mismatch")

    reasons = list(dict.fromkeys(reasons))
    verified = not reasons
    preferred_side = "none"
    relation = "not_comparable"
    confirmed_conflict = False
    if verified and left_number is not None and right_number is not None:
        equal = math.isclose(left_number, right_number, rel_tol=1e-9, abs_tol=1e-12)
        if mode == "replication":
            relation = "agreement" if equal else "reported_value_conflict"
            confirmed_conflict = not equal
        elif equal:
            relation = "tie"
            preferred_side = "tie"
        else:
            left_favorable = left_number > right_number if metric_direction == "higher" else left_number < right_number
            preferred_side = "left" if left_favorable else "right"
            relation = "left_more_favorable" if left_favorable else "right_more_favorable"

    left_name = _text((left_row or {}).get("paper") or (left_row or {}).get("source_name"), limit=240) or "Left source"
    right_name = _text((right_row or {}).get("paper") or (right_row or {}).get("source_name"), limit=240) or "Right source"
    metric_label = _text(dimensions.get("metric", {}).get("left_value"), limit=120) or "metric"
    dataset_label = _text(dimensions.get("dataset", {}).get("left_value"), limit=120) or "the stated dataset"
    if not verified:
        conclusion = "No comparative conclusion: " + ", ".join(reason.replace("_", " ") for reason in reasons[:8]) + "."
    elif mode == "replication" and confirmed_conflict:
        conclusion = (
            f"{left_name} reports {clean_spec['left_result']} and {right_name} reports {clean_spec['right_result']} "
            f"for {metric_label} on {dataset_label}. The reported values disagree under the audited matched contract; "
            "this is a reporting conflict, not proof of a broader scientific contradiction."
        )
    elif mode == "replication":
        conclusion = (
            f"{left_name} and {right_name} both report {clean_spec['left_result']} for {metric_label} on "
            f"{dataset_label} under the audited matched contract."
        )
    elif relation == "tie":
        conclusion = (
            f"{left_name} and {right_name} both report {clean_spec['left_result']} for {metric_label} on "
            f"{dataset_label}; the audited comparison is tied."
        )
    else:
        favored = left_name if preferred_side == "left" else right_name
        conclusion = (
            f"{left_name} reports {clean_spec['left_result']} and {right_name} reports {clean_spec['right_result']} "
            f"for {metric_label} on {dataset_label}; {favored} has the more favorable reported value because "
            f"{metric_direction} is better. This conclusion is limited to the audited contract and is not a general method ranking."
        )
    validation_ms = round((time.perf_counter() - validation_started) * 1000, 3)
    total_ms = round((time.perf_counter() - total_started) * 1000, 3)
    return {
        "id": audit_id,
        "contract_version": 1,
        "status": "verified" if verified else "not_comparable",
        "mode": mode,
        "input": clean_spec,
        "left_row_id": left_row_id,
        "right_row_id": right_row_id,
        "left_source_name": left_name,
        "right_source_name": right_name,
        "dimensions": dimension_audits,
        "metric": left_metric,
        "metric_direction": metric_direction,
        "result_unit": left_unit if left_unit == right_unit else "",
        "relation": relation,
        "preferred_side": preferred_side,
        "target_match_type": target_match_type,
        "confirmed_conflict": confirmed_conflict,
        "conclusion": conclusion,
        "reasons": reasons,
        "warnings": ["user_confirmed_mapping"] if user_confirmed_mappings else [],
        "user_confirmed_mappings": sorted(set(user_confirmed_mappings)),
        "evidence": evidence,
        "evidence_bindings": bindings,
        "phase_timings_ms": {
            "load_corpus": load_ms,
            "source_evidence_matching": matching_ms,
            "contract_validation": validation_ms,
            "total": total_ms,
        },
        "created_at": time.time(),
    }


def reaudit_evidence_comparisons(
    *,
    rows: list[dict[str, Any]],
    audits: list[dict[str, Any]],
    db_dir: str | Path,
) -> list[dict[str, Any]]:
    specs = [item.get("input") for item in audits if isinstance(item, dict) and isinstance(item.get("input"), dict)]
    if not specs:
        return []
    chunks = [item for item in load_all_chunks(Path(db_dir)) if isinstance(item, dict)]
    return [
        audit_evidence_comparison(rows=rows, spec=dict(spec), db_dir=db_dir, corpus_chunks=chunks)
        for spec in specs
    ]


def evidence_comparison_quality(audits: list[dict[str, Any]]) -> dict[str, Any]:
    active = [item for item in audits if isinstance(item, dict)]
    verified = [item for item in active if str(item.get("status") or "") == "verified"]
    conflicts = [item for item in verified if bool(item.get("confirmed_conflict"))]
    return {
        "comparison_audit_count": len(active),
        "verified_comparison_count": len(verified),
        "not_comparable_count": len(active) - len(verified),
        "confirmed_conflicts": [
            {
                "id": str(item.get("id") or ""),
                "conclusion": str(item.get("conclusion") or ""),
                "evidence_ids": [
                    str(evidence.get("id") or "")
                    for evidence in list(item.get("evidence") or [])
                    if isinstance(evidence, dict) and str(evidence.get("id") or "")
                ],
            }
            for item in conflicts
        ],
    }


def build_project_evidence_matrix(
    selected_items: list[dict[str, Any]],
    *,
    objective: str,
    db_dir: str | Path,
    existing_rows: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    context_items = list(research_brief_context(selected_items).get("items") or [])[:_MAX_MATRIX_SOURCES]
    chunks = [item for item in load_all_chunks(Path(db_dir)) if isinstance(item, dict)]
    prior_by_source = {
        _source_identity(row.get("source_path")): row
        for row in list(existing_rows or [])
        if isinstance(row, dict) and _source_identity(row.get("source_path"))
    }
    rows: list[dict[str, Any]] = []
    evidence: list[dict[str, Any]] = []
    evidence_seen: set[str] = set()
    for source in context_items:
        if not isinstance(source, dict):
            continue
        source_path = _text(source.get("sourcePath"), limit=1_200)
        source_name = _text(source.get("sourceName") or source.get("title"), limit=500)
        source_title = _text(source.get("title") or source_name, limit=800)
        source_key = source_path.replace("\\", "/").lower()
        row_id = f"row_{hashlib.sha1(source_key.encode('utf-8', errors='ignore')).hexdigest()[:16]}"
        previous = prior_by_source.get(_source_identity(source_path), {})
        row = {
            "id": row_id,
            "source_item_key": _text(source.get("key"), limit=500),
            "paper": source_title or source_name or Path(source_path).stem,
            "source_name": source_name or Path(source_path).name,
            "source_path": source_path,
            "authors": _text(source.get("authors"), limit=800),
            "year": _text(source.get("year"), limit=40),
            "doi": _text(source.get("doi"), limit=400),
            "notes": _text(previous.get("notes"), limit=4_000, multiline=True),
            "source_status": "active",
            "cells": {},
        }
        source_chunks = _source_chunks(chunks, source_path)
        source_retriever = BM25Retriever(source_chunks)
        used_quotes: set[str] = set()
        for field in MATRIX_CELL_FIELDS:
            selected = _best_cell_hit(
                source_chunks,
                field=field,
                objective=objective,
                retriever=source_retriever,
                excluded_quotes=used_quotes,
            )
            if not selected:
                row["cells"][field] = _empty_cell(field)
                continue
            hit, sentence = selected
            used_quotes.add(_normal(sentence))
            meta = _meta(hit)
            evidence_id = _evidence_id(source_path, field, hit, sentence)
            row["cells"][field] = {
                "field": field,
                "value": _text(sentence, limit=_MAX_CELL_VALUE),
                "support_status": "grounded",
                "evidence_ids": [evidence_id],
                "manual_override": False,
            }
            if evidence_id in evidence_seen:
                continue
            evidence_seen.add(evidence_id)
            evidence.append(
                {
                    "id": evidence_id,
                    "field": field,
                    "source_item_key": row["source_item_key"],
                    "source_path": source_path,
                    "source_name": row["source_name"],
                    "title": source_title,
                    "heading_path": _text(meta.get("heading_path") or meta.get("top_heading"), limit=800),
                    "location_label": _text(meta.get("location_label") or meta.get("heading_path"), limit=500),
                    "page_start": meta.get("page_start") or meta.get("page") or None,
                    "page_end": meta.get("page_end") or meta.get("page") or None,
                    "block_id": _text(meta.get("block_id"), limit=500),
                    "anchor_id": _text(meta.get("anchor_id") or meta.get("anchor"), limit=500),
                    "evidence_quote": _text(sentence, limit=_MAX_EVIDENCE_QUOTE),
                    "score": float(hit.get("score") or 0.0),
                }
            )
        rows.append(row)
    return rows, evidence, _comparison_flags(rows)


def evidence_matrix_cell_repair_candidates(
    matrix: dict[str, Any],
    gap: dict[str, Any],
    *,
    db_dir: str | Path,
    limit: int = 3,
    chunks: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Find strict, same-source repair candidates for one audited matrix cell."""
    if str(gap.get("kind") or "") not in {"missing_cell", "unsupported_cell"}:
        return []
    if str(gap.get("matrix_id") or "") != str(matrix.get("id") or ""):
        return []
    gap_revision = int(gap.get("matrix_revision") or 0)
    matrix_revision = int(matrix.get("revision") or 0)
    if gap_revision > 0 and gap_revision != matrix_revision:
        return []
    row_id = str(gap.get("row_id") or "")
    field = str(gap.get("field") or "")
    if field not in MATRIX_CELL_FIELDS or not row_id:
        return []
    row = next(
        (
            item
            for item in list(matrix.get("rows") or [])
            if isinstance(item, dict) and str(item.get("id") or "") == row_id
        ),
        None,
    )
    if not isinstance(row, dict):
        return []
    source_path = _text(row.get("source_path"), limit=1_200)
    source_identity = _source_identity(source_path)
    if not source_identity:
        return []
    corpus = [
        item
        for item in (chunks if chunks is not None else load_all_chunks(Path(db_dir)))
        if isinstance(item, dict)
    ]
    source_chunks = _source_chunks(corpus, source_path)
    if not source_chunks:
        return []

    query_terms, pattern = _FIELD_SPECS[field]
    objective = _text(matrix.get("objective"), limit=4_000)
    retriever = BM25Retriever(source_chunks)
    ranked: list[dict[str, Any]] = []
    seen_chunks: set[str] = set()
    for query in (query_terms, f"{objective} {query_terms}" if objective else ""):
        if not query:
            continue
        for hit in retriever.search(query, top_k=min(48, max(12, len(source_chunks)))):
            chunk_id = _text(hit.get("id") or _meta(hit).get("chunk_id"), limit=500)
            identity = chunk_id or str(id(hit))
            if identity in seen_chunks:
                continue
            seen_chunks.add(identity)
            ranked.append(hit)
    # Deep repair is an explicit action, so inspect remaining same-paper chunks
    # after BM25 ordering. Field-specific guards still reject weak sentences.
    for hit in source_chunks:
        chunk_id = _text(hit.get("id") or _meta(hit).get("chunk_id"), limit=500)
        identity = chunk_id or str(id(hit))
        if identity not in seen_chunks:
            seen_chunks.add(identity)
            ranked.append(hit)

    objective_tokens = {
        str(token).lower()
        for token in tokenize(objective)
        if len(str(token).strip()) >= 2
    }
    other_cell_values = {
        _normal(cell.get("value"))
        for other_field, cell in dict(row.get("cells") or {}).items()
        if other_field != field and isinstance(cell, dict) and _normal(cell.get("value"))
    }
    scored: list[tuple[float, str, dict[str, Any]]] = []
    seen_quotes: set[str] = set()
    for rank, hit in enumerate(ranked):
        meta = _meta(hit)
        if meta.get("evidence_ready") is False:
            continue
        if _source_identity(_source_path(hit)) != source_identity:
            continue
        heading = _text(meta.get("heading_path") or meta.get("top_heading"), limit=800)
        if _REFERENCE_HEADING_RE.search(heading):
            continue
        for sentence in _sentence_candidates(hit.get("text")):
            normalized = _normal(sentence)
            if not normalized or normalized in seen_quotes or normalized in other_cell_values:
                continue
            if _CAPTION_ONLY_RE.search(sentence):
                continue
            score = _candidate_score(
                sentence,
                field=field,
                pattern=pattern,
                objective_tokens=objective_tokens,
                heading=heading,
            )
            if score < 0:
                continue
            if pattern.search(heading):
                score += 0.6
            score += max(0.0, 0.35 - rank * 0.01)
            chunk_id = _text(hit.get("id") or meta.get("chunk_id"), limit=500)
            block_id = _text(meta.get("block_id"), limit=500)
            anchor_id = _text(meta.get("anchor_id") or meta.get("anchor"), limit=500)
            page_start = meta.get("page_start") or meta.get("page") or None
            page_end = meta.get("page_end") or meta.get("page") or None
            location_label = _text(meta.get("location_label") or heading, limit=500)
            if not (anchor_id or block_id or heading or page_start is not None):
                continue
            value = _text(sentence, limit=_MAX_CELL_VALUE)
            evidence_id = _evidence_id(source_path, field, hit, value)
            candidate_seed = "|".join(
                (
                    str(gap.get("gap_key") or ""),
                    str(matrix_revision),
                    row_id,
                    field,
                    evidence_id,
                )
            )
            candidate_id = f"repair_{hashlib.sha1(candidate_seed.encode('utf-8', errors='ignore')).hexdigest()[:18]}"
            seen_quotes.add(normalized)
            scored.append(
                (
                    score,
                    candidate_id,
                    {
                        "id": candidate_id,
                        "gap_id": str(gap.get("id") or ""),
                        "gap_key": str(gap.get("gap_key") or ""),
                        "matrix_id": str(matrix.get("id") or ""),
                        "matrix_revision": matrix_revision,
                        "row_id": row_id,
                        "field": field,
                        "value": value,
                        "source_path": source_path,
                        "source_name": _text(row.get("source_name") or row.get("paper"), limit=500),
                        "title": _text(row.get("paper") or row.get("source_name"), limit=800),
                        "chunk_id": chunk_id,
                        "evidence_id": evidence_id,
                        "evidence_quote": value,
                        "heading_path": heading,
                        "location_label": location_label,
                        "page_start": page_start,
                        "page_end": page_end,
                        "block_id": block_id,
                        "anchor_id": anchor_id,
                        "score": round(float(score), 6),
                        "same_source_verified": True,
                        "match_reason": "The field-specific extractor found this exact passage in the matrix row's source paper.",
                    },
                )
            )
    scored.sort(key=lambda item: (-item[0], item[1]))
    return [item[2] for item in scored[: max(1, min(8, int(limit or 3)))]]


def evidence_matrix_quality(
    *,
    rows: list[dict[str, Any]],
    evidence: list[dict[str, Any]],
    selected_items: list[dict[str, Any]],
    comparison_flags: list[dict[str, Any]] | None = None,
    comparison_audits: list[dict[str, Any]] | None = None,
) -> tuple[str, dict[str, Any]]:
    expected_items = list(research_brief_context(selected_items).get("items") or [])[:_MAX_MATRIX_SOURCES]
    expected_sources = {
        _source_identity(item.get("sourcePath"))
        for item in expected_items
        if isinstance(item, dict) and _source_identity(item.get("sourcePath"))
    }
    row_sources = {
        _source_identity(row.get("source_path"))
        for row in rows
        if isinstance(row, dict) and str(row.get("source_status") or "active") == "active"
    }
    evidence_by_id = {
        str(item.get("id") or ""): item
        for item in evidence
        if isinstance(item, dict) and str(item.get("id") or "")
    }
    missing_cells: list[dict[str, str]] = []
    unsupported_cells: list[dict[str, str]] = []
    populated_cells = 0
    supported_cells = 0
    source_evidence_counts: dict[str, int] = {source: 0 for source in expected_sources}
    for row in rows:
        if not isinstance(row, dict) or str(row.get("source_status") or "active") != "active":
            continue
        row_source = _source_identity(row.get("source_path"))
        cells = row.get("cells") if isinstance(row.get("cells"), dict) else {}
        for field in MATRIX_CELL_FIELDS:
            cell = cells.get(field) if isinstance(cells.get(field), dict) else {}
            value = _text(cell.get("value"), limit=_MAX_CELL_VALUE)
            if not value:
                missing_cells.append({"row_id": str(row.get("id") or ""), "field": field})
                continue
            populated_cells += 1
            evidence_ids = [str(item or "") for item in list(cell.get("evidence_ids") or []) if str(item or "")]
            matched = [evidence_by_id[item] for item in evidence_ids if item in evidence_by_id]
            supported = bool(matched) and all(
                _source_identity(item.get("source_path")) == row_source
                and _normal(value) in _normal(item.get("evidence_quote"))
                for item in matched
            )
            supported = supported and str(cell.get("support_status") or "") == "grounded"
            supported = supported and not bool(cell.get("manual_override"))
            if supported:
                supported_cells += 1
                source_evidence_counts[row_source] = source_evidence_counts.get(row_source, 0) + 1
            else:
                unsupported_cells.append({"row_id": str(row.get("id") or ""), "field": field})
    unexpected_sources = sorted(source for source in row_sources if source and source not in expected_sources)
    missing_sources = sorted(source for source in expected_sources if source not in row_sources)
    sources_without_evidence = sorted(
        source for source in expected_sources if source_evidence_counts.get(source, 0) <= 0
    )
    unexpected_evidence = sorted(
        {
            _source_identity(item.get("source_path"))
            for item in evidence
            if isinstance(item, dict)
            and _source_identity(item.get("source_path"))
            and _source_identity(item.get("source_path")) not in expected_sources
        }
    )
    reasons: list[str] = []
    if not rows:
        reasons.append("no_rows")
    if missing_sources:
        reasons.append("selected_sources_without_rows")
    if sources_without_evidence:
        reasons.append("selected_sources_without_evidence")
    if unexpected_sources or unexpected_evidence:
        reasons.append("unexpected_sources")
    if unsupported_cells:
        reasons.append("unsupported_cells")
    if populated_cells <= 0:
        reasons.append("no_supported_cells")
    completeness = (
        populated_cells / max(1, len(expected_sources) * len(MATRIX_CELL_FIELDS))
        if expected_sources
        else 0.0
    )
    flags = [item for item in list(comparison_flags or []) if isinstance(item, dict)]
    comparison_summary = evidence_comparison_quality(
        [item for item in list(comparison_audits or []) if isinstance(item, dict)]
    )
    quality = {
        "contract_version": 2,
        "generation_mode": "extractive",
        "selected_source_count": len(expected_sources),
        "row_count": len(rows),
        "covered_source_count": len(expected_sources) - len(sources_without_evidence),
        "populated_cell_count": populated_cells,
        "supported_cell_count": supported_cells,
        "unsupported_cell_count": len(unsupported_cells),
        "missing_cell_count": len(missing_cells),
        "completeness": round(max(0.0, min(1.0, completeness)), 4),
        "missing_cells": missing_cells[:80],
        "unsupported_cells": unsupported_cells[:80],
        "missing_sources": missing_sources,
        "sources_without_evidence": sources_without_evidence,
        "unexpected_sources": sorted(set(unexpected_sources + unexpected_evidence)),
        "comparison_flags": flags,
        **comparison_summary,
        "reasons": reasons,
        "warnings": ["missing_cells"] if missing_cells else [],
        "edited_after_verification": False,
    }
    return ("verified" if not reasons else "needs_review"), quality


def apply_evidence_matrix_cell_repair(
    matrix: dict[str, Any],
    gap: dict[str, Any],
    candidate: dict[str, Any],
    *,
    db_dir: str | Path,
) -> dict[str, Any]:
    """Apply a recomputed same-source candidate and rerun matrix contracts."""
    matrix_id = str(matrix.get("id") or "")
    matrix_revision = int(matrix.get("revision") or 0)
    row_id = str(gap.get("row_id") or "")
    field = str(gap.get("field") or "")
    if (
        str(candidate.get("gap_key") or "") != str(gap.get("gap_key") or "")
        or str(candidate.get("matrix_id") or "") != matrix_id
        or int(candidate.get("matrix_revision") or 0) != matrix_revision
        or str(candidate.get("row_id") or "") != row_id
        or str(candidate.get("field") or "") != field
        or field not in MATRIX_CELL_FIELDS
    ):
        raise ValueError("repair candidate does not match the current matrix gap")
    rows = copy.deepcopy([item for item in list(matrix.get("rows") or []) if isinstance(item, dict)])
    row = next((item for item in rows if str(item.get("id") or "") == row_id), None)
    if not isinstance(row, dict):
        raise ValueError("repair row is unavailable")
    if (
        not bool(candidate.get("same_source_verified"))
        or _source_identity(candidate.get("source_path")) != _source_identity(row.get("source_path"))
    ):
        raise ValueError("repair evidence must come from the matrix row's source paper")
    value = _text(candidate.get("value"), limit=_MAX_CELL_VALUE)
    quote = _text(candidate.get("evidence_quote"), limit=_MAX_EVIDENCE_QUOTE)
    if not value or _normal(value) not in _normal(quote):
        raise ValueError("repair value is not bound to its exact evidence quote")
    evidence_id = _text(candidate.get("evidence_id"), limit=300)
    if not evidence_id:
        raise ValueError("repair evidence identity is missing")

    cells = copy.deepcopy(dict(row.get("cells") or {}))
    previous_cell = dict(cells.get(field) or {}) if isinstance(cells.get(field), dict) else {}
    previous_evidence_ids = {
        str(item or "") for item in list(previous_cell.get("evidence_ids") or []) if str(item or "")
    }
    cells[field] = {
        "field": field,
        "value": value,
        "support_status": "grounded",
        "evidence_ids": [evidence_id],
        "manual_override": False,
        "repair_confirmed": True,
    }
    row["cells"] = cells

    referenced_elsewhere = {
        str(evidence_ref or "")
        for matrix_row in rows
        if isinstance(matrix_row, dict)
        for other_field, cell in dict(matrix_row.get("cells") or {}).items()
        if isinstance(cell, dict) and not (str(matrix_row.get("id") or "") == row_id and other_field == field)
        for evidence_ref in list(cell.get("evidence_ids") or [])
        if str(evidence_ref or "")
    }
    evidence = [
        copy.deepcopy(item)
        for item in list(matrix.get("evidence") or [])
        if isinstance(item, dict)
        and str(item.get("id") or "") != evidence_id
        and not (
            str(item.get("id") or "") in previous_evidence_ids
            and str(item.get("id") or "") not in referenced_elsewhere
        )
    ]
    evidence.append(
        {
            "id": evidence_id,
            "field": field,
            "source_item_key": _text(row.get("source_item_key"), limit=500),
            "source_path": _text(row.get("source_path"), limit=1_200),
            "source_name": _text(row.get("source_name") or row.get("paper"), limit=500),
            "title": _text(row.get("paper") or row.get("source_name"), limit=800),
            "heading_path": _text(candidate.get("heading_path"), limit=800),
            "location_label": _text(candidate.get("location_label"), limit=500),
            "page_start": candidate.get("page_start"),
            "page_end": candidate.get("page_end"),
            "block_id": _text(candidate.get("block_id"), limit=500),
            "anchor_id": _text(candidate.get("anchor_id"), limit=500),
            "evidence_quote": quote,
            "chunk_id": _text(candidate.get("chunk_id"), limit=500),
            "score": float(candidate.get("score") or 0.0),
            "repair_gap_id": str(gap.get("id") or ""),
            "repair_gap_key": str(gap.get("gap_key") or ""),
        }
    )

    current_audits = copy.deepcopy(
        [item for item in list(matrix.get("comparison_audits") or []) if isinstance(item, dict)]
    )
    positions = [
        index
        for index, audit in enumerate(current_audits)
        if str(audit.get("left_row_id") or "") == row_id or str(audit.get("right_row_id") or "") == row_id
    ]
    refreshed = reaudit_evidence_comparisons(
        rows=rows,
        audits=[current_audits[index] for index in positions],
        db_dir=db_dir,
    )
    comparison_audits = list(current_audits)
    for index, audit in zip(positions, refreshed):
        comparison_audits[index] = audit
    comparison_flags = evidence_matrix_comparison_flags(rows)
    selected_items = [
        item for item in list(matrix.get("source_items") or []) if isinstance(item, dict)
    ]
    quality_status, quality = evidence_matrix_quality(
        rows=rows,
        evidence=evidence,
        selected_items=selected_items,
        comparison_flags=comparison_flags,
        comparison_audits=comparison_audits,
    )
    current_quality = matrix.get("quality") if isinstance(matrix.get("quality"), dict) else {}
    if isinstance(current_quality.get("source_watch_snapshot"), dict):
        quality["source_watch_snapshot"] = copy.deepcopy(current_quality["source_watch_snapshot"])
    quality["last_research_gap_repair"] = {
        "contract_version": 1,
        "gap_id": str(gap.get("id") or ""),
        "gap_key": str(gap.get("gap_key") or ""),
        "row_id": row_id,
        "field": field,
        "source_path": _text(row.get("source_path"), limit=1_200),
        "evidence_id": evidence_id,
        "candidate_id": str(candidate.get("id") or ""),
        "reaudited_comparison_count": len(refreshed),
    }
    return {
        "rows": rows,
        "evidence": evidence,
        "comparison_flags": comparison_flags,
        "comparison_audits": comparison_audits,
        "quality_status": quality_status,
        "quality": quality,
        "reaudited_comparison_count": len(refreshed),
    }


def evidence_matrix_hits(record: dict[str, Any], *, limit: int = 20) -> list[dict[str, Any]]:
    rows = [item for item in list(record.get("rows") or []) if isinstance(item, dict)]
    evidence = [item for item in list(record.get("evidence") or []) if isinstance(item, dict)]
    evidence_by_id = {str(item.get("id") or ""): item for item in evidence if str(item.get("id") or "")}
    comparison_hits: list[dict[str, Any]] = []
    for audit in list(record.get("comparison_audits") or []):
        if not isinstance(audit, dict) or str(audit.get("status") or "") != "verified":
            continue
        spec = audit.get("input") if isinstance(audit.get("input"), dict) else {}
        dimensions = _comparison_dimensions(spec)
        direction = _text(audit.get("metric_direction"), limit=40)
        audit_evidence = {
            str(item.get("id") or ""): item
            for item in list(audit.get("evidence") or [])
            if isinstance(item, dict) and str(item.get("id") or "")
        }
        bindings = audit.get("evidence_bindings") if isinstance(audit.get("evidence_bindings"), dict) else {}
        for side in ("left", "right"):
            side_bindings = bindings.get(side) if isinstance(bindings.get(side), dict) else {}
            item = audit_evidence.get(str(side_bindings.get("result") or ""))
            dataset = _text(dimensions.get("dataset", {}).get(f"{side}_value"), limit=160)
            metric = _text(dimensions.get("metric", {}).get(f"{side}_value"), limit=120)
            target = _text(spec.get(f"{side}_target"), limit=240)
            result = _text(spec.get(f"{side}_result"), limit=80)
            if not item or not dataset or not metric or not target or not result:
                continue
            observation = f"{dataset} {metric} ({direction} is better): {target} = {result}."
            comparison_hits.append(
                {
                    "id": str(item.get("id") or ""),
                    "text": observation,
                    "score": 10.0,
                    "meta": {
                        "source_path": _text(item.get("source_path"), limit=1_200),
                        "source_name": _text(item.get("source_name"), limit=500),
                        "title": _text(item.get("title"), limit=800),
                        "heading_path": _text(item.get("heading_path"), limit=800),
                        "location_label": _text(item.get("location_label"), limit=500),
                        "page_start": item.get("page_start"),
                        "page_end": item.get("page_end"),
                        "block_id": _text(item.get("block_id"), limit=500),
                        "anchor_id": _text(item.get("anchor_id"), limit=500),
                        "matrix_field": "comparison_result",
                        "comparison_audit_id": _text(audit.get("id"), limit=120),
                        "comparison_relation": _text(audit.get("relation"), limit=120),
                        "comparison_source_quote": _text(item.get("evidence_quote"), limit=_MAX_EVIDENCE_QUOTE, multiline=True),
                    },
                }
            )
    comparison_hits = comparison_hits[: max(0, int(limit))]
    cell_limit = max(0, int(limit) - len(comparison_hits))
    selected: list[dict[str, Any]] = []
    seen_evidence: set[str] = set()
    for field in MATRIX_CELL_FIELDS if cell_limit > 0 else ():
        for row in rows:
            cells = row.get("cells") if isinstance(row.get("cells"), dict) else {}
            cell = cells.get(field) if isinstance(cells.get(field), dict) else {}
            if str(cell.get("support_status") or "") != "grounded" or bool(cell.get("manual_override")):
                continue
            for evidence_id in list(cell.get("evidence_ids") or []):
                key = str(evidence_id or "")
                item = evidence_by_id.get(key)
                if not item or key in seen_evidence:
                    continue
                seen_evidence.add(key)
                selected.append(item)
                break
            if len(selected) >= cell_limit:
                break
        if len(selected) >= cell_limit:
            break
    hits: list[dict[str, Any]] = []
    for item in selected[: max(1, int(limit))]:
        hits.append(
            {
                "id": str(item.get("id") or ""),
                "text": _text(item.get("evidence_quote"), limit=_MAX_EVIDENCE_QUOTE),
                "score": float(item.get("score") or 0.0),
                "meta": {
                    "source_path": _text(item.get("source_path"), limit=1_200),
                    "source_name": _text(item.get("source_name"), limit=500),
                    "title": _text(item.get("title"), limit=800),
                    "heading_path": _text(item.get("heading_path"), limit=800),
                    "location_label": _text(item.get("location_label"), limit=500),
                    "page_start": item.get("page_start"),
                    "page_end": item.get("page_end"),
                    "block_id": _text(item.get("block_id"), limit=500),
                    "anchor_id": _text(item.get("anchor_id"), limit=500),
                    "matrix_field": _text(item.get("field"), limit=80),
                },
            }
        )
    return [*hits, *comparison_hits][: max(1, int(limit))]


def _cell_value(row: dict[str, Any], field: str) -> str:
    cells = row.get("cells") if isinstance(row.get("cells"), dict) else {}
    cell = cells.get(field) if isinstance(cells.get(field), dict) else {}
    return _text(cell.get("value"), limit=_MAX_CELL_VALUE)


def _tabular_value(value: object, *, limit: int) -> str:
    text = _text(value, limit=limit, multiline=True)
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def evidence_matrix_markdown(record: dict[str, Any]) -> str:
    title = _text(record.get("title"), limit=240) or "Evidence matrix"
    objective = _text(record.get("objective"), limit=4_000, multiline=True)
    status = _text(record.get("quality_status"), limit=40) or "draft"
    revision = max(1, int(record.get("revision") or 1))
    lines = [f"# {title}", "", f"> Evidence status: {status}; revision: {revision}."]
    if objective:
        lines.extend(["", "## Research objective", "", objective])
    headers = ["Paper", "Method", "Dataset / experiment", "Metric", "Key result", "Limitation", "Notes"]
    lines.extend(["", "## Matrix", "", "| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"])
    for row in list(record.get("rows") or []):
        if not isinstance(row, dict):
            continue
        values = [
            _text(row.get("paper") or row.get("source_name"), limit=500),
            *[_cell_value(row, field) for field in MATRIX_CELL_FIELDS],
            _text(row.get("notes"), limit=4_000, multiline=True),
        ]
        escaped = [value.replace("|", "\\|").replace("\n", "<br>") or "—" for value in values]
        lines.append("| " + " | ".join(escaped) + " |")
    comparisons = [item for item in list(record.get("comparison_audits") or []) if isinstance(item, dict)]
    if comparisons:
        lines.extend(["", "## Comparison audits"])
        for item in comparisons:
            status = _text(item.get("status"), limit=40) or "not_comparable"
            mode = _text(item.get("mode"), limit=40) or "ranking"
            lines.extend(
                [
                    "",
                    f"### {_text(item.get('left_source_name'), limit=240)} / {_text(item.get('right_source_name'), limit=240)}",
                    "",
                    f"- Status: {status}",
                    f"- Mode: {mode}",
                    f"- Conclusion: {_text(item.get('conclusion'), limit=1_600, multiline=True)}",
                ]
            )
            reasons = [_text(reason, limit=120) for reason in list(item.get("reasons") or []) if _text(reason, limit=120)]
            if reasons:
                lines.append(f"- Boundaries: {', '.join(reasons)}")
    evidence = [item for item in list(record.get("evidence") or []) if isinstance(item, dict)]
    if evidence:
        lines.extend(["", "## Evidence appendix"])
        for item in evidence:
            label = _text(item.get("source_name") or item.get("source_path"), limit=500) or "Source"
            field = _text(item.get("field"), limit=80)
            locator = _text(item.get("heading_path") or item.get("location_label"), limit=800)
            lines.extend(["", f"### {label} — {field}{f' — {locator}' if locator else ''}"])
            quote = _text(item.get("evidence_quote"), limit=_MAX_EVIDENCE_QUOTE, multiline=True)
            if quote:
                lines.extend(["", f"> {quote.replace(chr(10), chr(10) + '> ')}"])
    return "\n".join(lines).strip() + "\n"


def evidence_matrix_csv(record: dict[str, Any]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(
        [
            "paper",
            "source_path",
            "method",
            "dataset_or_experiment",
            "metric",
            "key_result",
            "limitation",
            "notes",
        ]
    )
    for row in list(record.get("rows") or []):
        if not isinstance(row, dict):
            continue
        writer.writerow(
            [
                _tabular_value(row.get("paper") or row.get("source_name"), limit=500),
                _tabular_value(row.get("source_path"), limit=1_200),
                *[_tabular_value(_cell_value(row, field), limit=_MAX_CELL_VALUE) for field in MATRIX_CELL_FIELDS],
                _tabular_value(row.get("notes"), limit=4_000),
            ]
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def evidence_matrix_xlsx(record: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence Matrix"
    headers = [
        "Paper",
        "Source path",
        "Method",
        "Dataset / experiment",
        "Metric",
        "Key result",
        "Limitation",
        "Notes",
    ]
    sheet.append(headers)
    for row in list(record.get("rows") or []):
        if not isinstance(row, dict):
            continue
        sheet.append(
            [
                _tabular_value(row.get("paper") or row.get("source_name"), limit=500),
                _tabular_value(row.get("source_path"), limit=1_200),
                *[_tabular_value(_cell_value(row, field), limit=_MAX_CELL_VALUE) for field in MATRIX_CELL_FIELDS],
                _tabular_value(row.get("notes"), limit=4_000),
            ]
        )
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [28, 42, 48, 42, 34, 48, 42, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    evidence_sheet = workbook.create_sheet("Evidence")
    evidence_headers = ["ID", "Paper", "Field", "Locator", "Page", "Evidence quote", "Source path"]
    evidence_sheet.append(evidence_headers)
    for item in list(record.get("evidence") or []):
        if not isinstance(item, dict):
            continue
        evidence_sheet.append(
            [
                _tabular_value(item.get("id"), limit=120),
                _tabular_value(item.get("source_name"), limit=500),
                _tabular_value(item.get("field"), limit=80),
                _tabular_value(item.get("heading_path") or item.get("location_label"), limit=800),
                item.get("page_start") or "",
                _tabular_value(item.get("evidence_quote"), limit=_MAX_EVIDENCE_QUOTE),
                _tabular_value(item.get("source_path"), limit=1_200),
            ]
        )
    for cell in evidence_sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for index, width in enumerate([22, 28, 18, 36, 10, 72, 42], start=1):
        evidence_sheet.column_dimensions[chr(64 + index)].width = width
    for row in evidence_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    evidence_sheet.freeze_panes = "A2"
    evidence_sheet.auto_filter.ref = evidence_sheet.dimensions
    comparisons = [item for item in list(record.get("comparison_audits") or []) if isinstance(item, dict)]
    if comparisons:
        comparison_sheet = workbook.create_sheet("Comparison Audits")
        comparison_sheet.append(
            ["ID", "Status", "Mode", "Left source", "Right source", "Metric", "Relation", "Conclusion", "Boundaries"]
        )
        for item in comparisons:
            comparison_sheet.append(
                [
                    _tabular_value(item.get("id"), limit=120),
                    _tabular_value(item.get("status"), limit=80),
                    _tabular_value(item.get("mode"), limit=80),
                    _tabular_value(item.get("left_source_name"), limit=500),
                    _tabular_value(item.get("right_source_name"), limit=500),
                    _tabular_value(item.get("metric"), limit=120),
                    _tabular_value(item.get("relation"), limit=120),
                    _tabular_value(item.get("conclusion"), limit=1_600),
                    _tabular_value(", ".join(str(reason) for reason in list(item.get("reasons") or [])), limit=1_200),
                ]
            )
        for cell in comparison_sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        for index, width in enumerate([22, 18, 16, 28, 28, 16, 24, 72, 48], start=1):
            comparison_sheet.column_dimensions[chr(64 + index)].width = width
        for row in comparison_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        comparison_sheet.freeze_panes = "A2"
        comparison_sheet.auto_filter.ref = comparison_sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
