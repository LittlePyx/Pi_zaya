from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from kb.evidence_matrix import (
    MATRIX_CELL_FIELDS,
    audit_evidence_comparison,
    build_project_evidence_matrix,
    evidence_comparison_quality,
    evidence_matrix_csv,
    evidence_matrix_hits,
    evidence_matrix_markdown,
    evidence_matrix_quality,
    evidence_matrix_xlsx,
)


def _chunk(source: str, heading: str, text: str, index: int) -> dict:
    return {
        "id": f"chunk-{index}",
        "text": text,
        "meta": {
            "source_path": source,
            "source_name": Path(source).stem,
            "heading_path": heading,
            "page": index,
            "block_id": f"block-{index}",
        },
    }


def test_build_matrix_is_source_balanced_grounded_and_preserves_notes(monkeypatch) -> None:
    sources = ["F:/papers/a.md", "F:/papers/b.md"]
    chunks = []
    for source_index, source in enumerate(sources, start=1):
        chunks.extend(
            [
                _chunk(source, "Method", f"The method uses a coded optical network for source {source_index}.", source_index * 10 + 1),
                _chunk(source, "Experiments", f"Experiments use the Dataset-{source_index} benchmark with 100 samples.", source_index * 10 + 2),
                _chunk(source, "Metrics", f"The evaluation metric is PSNR and reaches {30 + source_index} dB.", source_index * 10 + 3),
                _chunk(source, "Results", f"Results improve reconstruction performance by {source_index + 2} percent.", source_index * 10 + 4),
                _chunk(source, "Limitations", f"A limitation remains under low-light conditions for source {source_index}.", source_index * 10 + 5),
            ]
        )
    monkeypatch.setattr("kb.evidence_matrix.load_all_chunks", lambda _db_dir: chunks)
    selected = [
        {"key": "a", "title": "Paper A", "sourcePath": sources[0]},
        {"key": "b", "title": "Paper B", "sourcePath": sources[1]},
    ]
    previous = [{"source_path": sources[0], "notes": "Keep this manual note."}]

    rows, evidence, flags = build_project_evidence_matrix(
        selected,
        objective="Compare reconstruction quality.",
        db_dir="unused",
        existing_rows=previous,
    )

    assert len(rows) == 2
    assert rows[0]["notes"] == "Keep this manual note."
    assert all(set(row["cells"]) == set(MATRIX_CELL_FIELDS) for row in rows)
    assert all(row["cells"][field]["support_status"] == "grounded" for row in rows for field in MATRIX_CELL_FIELDS)
    assert {item["source_path"] for item in evidence} == set(sources)
    assert any(item["code"] == "experimental_conditions_differ" for item in flags)

    status, quality = evidence_matrix_quality(
        rows=rows,
        evidence=evidence,
        selected_items=selected,
        comparison_flags=flags,
    )
    assert status == "verified"
    assert quality["covered_source_count"] == 2
    assert quality["supported_cell_count"] == quality["populated_cell_count"] == 10
    assert quality["unexpected_sources"] == []

    record = {"rows": rows, "evidence": evidence}
    hits = evidence_matrix_hits(record)
    assert {_hit["meta"]["source_path"] for _hit in hits[:2]} == set(sources)


def test_matrix_quality_rejects_manual_or_cross_source_cell_support() -> None:
    selected = [{"key": "a", "title": "Paper A", "sourcePath": "F:/papers/a.md"}]
    rows = [
        {
            "id": "row-a",
            "source_path": "F:/papers/a.md",
            "source_status": "active",
            "cells": {
                "method": {
                    "value": "A manually rewritten method claim.",
                    "support_status": "needs_review",
                    "evidence_ids": ["ev-1"],
                    "manual_override": True,
                }
            },
        }
    ]
    evidence = [
        {
            "id": "ev-1",
            "source_path": "F:/papers/b.md",
            "evidence_quote": "A different paper reports a method.",
        }
    ]

    status, quality = evidence_matrix_quality(
        rows=rows,
        evidence=evidence,
        selected_items=selected,
    )

    assert status == "needs_review"
    assert "unsupported_cells" in quality["reasons"]
    assert "unexpected_sources" in quality["reasons"]


def test_build_matrix_leaves_ambiguous_reference_and_positive_contrast_cells_empty(monkeypatch) -> None:
    source = "F:/papers/conservative.md"
    chunks = [
        _chunk(source, "References", "Ichioka, Thin observation module: experimental verification with 100 samples.", 1),
        _chunk(source, "Introduction", "Conventional methods employ an iterative network for reconstruction.", 2),
        _chunk(source, "Results", "Although simplified, its performance is equal to or better than baseline.", 3),
        _chunk(source, "Experiments", "Figure 7. Experimental setup of the system.", 4),
        _chunk(source, "Results", "The results are computed from images produced by state-of-the-art methods.", 5),
        _chunk(source, "Results", "Figure 4C shows examples of difference map stacks.", 6),
        _chunk(source, "Results", "Previous winning solution on the Image Deblurring Challenge Track2.", 7),
    ]
    monkeypatch.setattr("kb.evidence_matrix.load_all_chunks", lambda _db_dir: chunks)

    rows, evidence, _flags = build_project_evidence_matrix(
        [{"key": "paper", "title": "Paper", "sourcePath": source}],
        objective="Compare the proposed method and evidence.",
        db_dir="unused",
    )

    assert rows[0]["cells"]["method"]["support_status"] == "missing"
    assert rows[0]["cells"]["dataset_or_experiment"]["support_status"] == "missing"
    assert rows[0]["cells"]["metric"]["support_status"] == "missing"
    assert rows[0]["cells"]["key_result"]["support_status"] == "grounded"
    assert rows[0]["cells"]["limitation"]["support_status"] == "missing"
    assert len(evidence) == 1
    assert evidence[0]["field"] == "key_result"


def test_matrix_exports_include_matrix_and_evidence() -> None:
    row = {
        "paper": "Paper A",
        "source_name": "Paper A",
        "source_path": "F:/papers/a.md",
        "notes": "Review Figure 2.",
        "cells": {
            field: {"value": f"{field} evidence", "support_status": "grounded", "evidence_ids": [f"ev-{field}"]}
            for field in MATRIX_CELL_FIELDS
        },
    }
    record = {
        "title": "Imaging evidence matrix",
        "objective": "Compare methods.",
        "quality_status": "verified",
        "revision": 2,
        "rows": [row],
        "evidence": [
            {
                "id": "ev-method",
                "source_name": "Paper A",
                "source_path": "F:/papers/a.md",
                "field": "method",
                "heading_path": "Method / Architecture",
                "page_start": 3,
                "evidence_quote": "The method uses a coded optical network.",
            }
        ],
    }

    markdown = evidence_matrix_markdown(record)
    assert "Imaging evidence matrix" in markdown
    assert "Evidence appendix" in markdown
    assert "Method / Architecture" in markdown
    csv_payload = evidence_matrix_csv(record).decode("utf-8-sig")
    assert "Paper A" in csv_payload
    assert "Review Figure 2" in csv_payload

    workbook = load_workbook(BytesIO(evidence_matrix_xlsx(record)))
    assert workbook.sheetnames == ["Evidence Matrix", "Evidence"]
    assert workbook["Evidence Matrix"]["A2"].value == "Paper A"
    assert workbook["Evidence"]["F2"].value == "The method uses a coded optical network."


def test_matrix_tabular_exports_escape_spreadsheet_formulas() -> None:
    row = {
        "paper": "Paper A",
        "source_path": "F:/papers/a.md",
        "notes": "=HYPERLINK(\"https://example.invalid\")",
        "cells": {field: {"value": ""} for field in MATRIX_CELL_FIELDS},
    }
    record = {"rows": [row], "evidence": []}

    assert "'=HYPERLINK" in evidence_matrix_csv(record).decode("utf-8-sig")
    workbook = load_workbook(BytesIO(evidence_matrix_xlsx(record)))
    assert workbook["Evidence Matrix"]["H2"].value.startswith("'=HYPERLINK")


def _comparison_rows() -> list[dict]:
    return [
        {
            "id": "row-left",
            "source_item_key": "left",
            "paper": "Paper Left",
            "source_name": "Paper Left",
            "source_path": "F:/papers/left.md",
            "source_status": "active",
            "cells": {},
        },
        {
            "id": "row-right",
            "source_item_key": "right",
            "paper": "Paper Right",
            "source_name": "Paper Right",
            "source_path": "F:/papers/right.md",
            "source_status": "active",
            "cells": {},
        },
    ]


def _comparison_spec(*, mode: str = "ranking", right_metric: str = "LPIPS", right_result: str = ".0445") -> dict:
    return {
        "mode": mode,
        "left_row_id": "row-left",
        "right_row_id": "row-right",
        "dimensions": [
            {"dimension": "task", "left_value": "SCI image reconstruction", "right_value": "SCI image reconstruction"},
            {"dimension": "dataset", "left_value": "Cozy2room", "right_value": "Cozy2room"},
            {
                "dimension": "evaluation_protocol",
                "left_value": "static datasets",
                "right_value": "synthetic datasets",
                "mapping_confirmed": True,
            },
            {"dimension": "metric", "left_value": "LPIPS", "right_value": right_metric},
        ],
        "left_target": "SCIGS(ours)",
        "right_target": "ours",
        "target_mapping_confirmed": mode == "replication",
        "left_result": ".0423",
        "right_result": right_result,
    }


def _comparison_chunks() -> list[dict]:
    return [
        _chunk(
            "F:/papers/left.md",
            "Experiments / Table 1",
            "Quantitative SCI image reconstruction comparisons on the static datasets. "
            "Cozy2room LPIPS ↓ (lower is better): SCIGS(ours) = .0423",
            41,
        ),
        _chunk(
            "F:/papers/right.md",
            "Experiments / Table 1",
            "Quantitative SCI image reconstruction comparisons on the synthetic datasets. "
            "Cozy2room LPIPS ↓ (lower is better): ours = .0445",
            42,
        ),
    ]


def test_comparison_audit_verifies_only_joint_same_source_quantitative_evidence(monkeypatch) -> None:
    monkeypatch.setattr("kb.evidence_matrix.load_all_chunks", lambda _db_dir: _comparison_chunks())

    audit = audit_evidence_comparison(rows=_comparison_rows(), spec=_comparison_spec(), db_dir="unused")

    assert audit["status"] == "verified"
    assert audit["metric"] == "lpips"
    assert audit["metric_direction"] == "lower"
    assert audit["relation"] == "left_more_favorable"
    assert audit["preferred_side"] == "left"
    assert audit["confirmed_conflict"] is False
    assert audit["user_confirmed_mappings"] == ["evaluation_protocol"]
    assert {item["side"] for item in audit["evidence"]} == {"left", "right"}
    assert all(item["source_path"].endswith(f"{item['side']}.md") for item in audit["evidence"])
    assert "not a general method ranking" in audit["conclusion"]


def test_comparison_audit_refuses_metric_mismatch_and_fabricated_result(monkeypatch) -> None:
    chunks = _comparison_chunks()
    chunks[1]["text"] = chunks[1]["text"].replace("LPIPS", "SSIM")
    monkeypatch.setattr("kb.evidence_matrix.load_all_chunks", lambda _db_dir: chunks)

    mismatch = audit_evidence_comparison(
        rows=_comparison_rows(),
        spec=_comparison_spec(right_metric="SSIM"),
        db_dir="unused",
    )
    fabricated = audit_evidence_comparison(
        rows=_comparison_rows(),
        spec=_comparison_spec(right_result=".9999"),
        db_dir="unused",
    )

    assert mismatch["status"] == "not_comparable"
    assert "metric_mismatch" in mismatch["reasons"]
    assert "unsupported_or_mismatched_metric" in mismatch["reasons"]
    assert fabricated["status"] == "not_comparable"
    assert "right_result_evidence_missing" in fabricated["reasons"]
    assert fabricated["preferred_side"] == "none"


def test_replication_audit_records_only_a_verified_reporting_conflict(monkeypatch) -> None:
    monkeypatch.setattr("kb.evidence_matrix.load_all_chunks", lambda _db_dir: _comparison_chunks())
    spec = _comparison_spec(mode="replication")
    spec["left_target"] = "SCINeRF"
    spec["right_target"] = "ours"
    chunks = _comparison_chunks()
    chunks[0]["text"] = chunks[0]["text"].replace("SCIGS(ours)", "SCINeRF")
    monkeypatch.setattr("kb.evidence_matrix.load_all_chunks", lambda _db_dir: chunks)

    audit = audit_evidence_comparison(rows=_comparison_rows(), spec=spec, db_dir="unused")
    summary = evidence_comparison_quality([audit])

    assert audit["status"] == "verified"
    assert audit["relation"] == "reported_value_conflict"
    assert audit["confirmed_conflict"] is True
    assert audit["target_match_type"] == "user_confirmed"
    assert summary["confirmed_conflicts"][0]["id"] == audit["id"]


def test_verified_comparison_adds_source_specific_brief_hits_without_adding_a_cross_source_claim(monkeypatch) -> None:
    monkeypatch.setattr("kb.evidence_matrix.load_all_chunks", lambda _db_dir: _comparison_chunks())
    rows = _comparison_rows()
    audit = audit_evidence_comparison(rows=rows, spec=_comparison_spec(), db_dir="unused")

    hits = evidence_matrix_hits({"rows": rows, "evidence": [], "comparison_audits": [audit]})

    assert len(hits) == 2
    assert {hit["meta"]["source_path"] for hit in hits} == {"F:/papers/left.md", "F:/papers/right.md"}
    assert all(hit["meta"]["matrix_field"] == "comparison_result" for hit in hits)
    assert any("SCIGS(ours) = .0423" in hit["text"] for hit in hits)
    assert any("ours = .0445" in hit["text"] for hit in hits)
    assert all("more favorable" not in hit["text"] for hit in hits)


def test_matrix_brief_hits_reserve_every_source_before_dense_comparisons() -> None:
    rows = []
    evidence = []
    for index in range(3):
        evidence_id = f"cell-{index}"
        source_path = f"F:/papers/source-{index}.md"
        rows.append(
            {
                "id": f"row-{index}",
                "source_path": source_path,
                "source_name": f"Paper {index}",
                "source_status": "active",
                "cells": {
                    "method": {
                        "support_status": "grounded",
                        "manual_override": False,
                        "evidence_ids": [evidence_id],
                    }
                },
            }
        )
        evidence.append(
            {
                "id": evidence_id,
                "field": "method",
                "source_path": source_path,
                "source_name": f"Paper {index}",
                "evidence_quote": f"Paper {index} reports its own grounded method.",
            }
        )
    comparison_audits = []
    for index in range(12):
        comparison_audits.append(
            {
                "id": f"comparison-{index}",
                "status": "verified",
                "metric_direction": "lower",
                "input": {
                    "dimensions": [
                        {"dimension": "dataset", "left_value": f"Dataset {index}", "right_value": f"Dataset {index}"},
                        {"dimension": "metric", "left_value": "LPIPS", "right_value": "LPIPS"},
                    ],
                    "left_target": "Paper 0",
                    "right_target": "Paper 1",
                    "left_result": ".10",
                    "right_result": ".20",
                },
                "evidence": [
                    {"id": f"left-{index}", "source_path": "F:/papers/source-0.md", "source_name": "Paper 0"},
                    {"id": f"right-{index}", "source_path": "F:/papers/source-1.md", "source_name": "Paper 1"},
                ],
                "evidence_bindings": {
                    "left": {"result": f"left-{index}"},
                    "right": {"result": f"right-{index}"},
                },
            }
        )

    hits = evidence_matrix_hits(
        {"rows": rows, "evidence": evidence, "comparison_audits": comparison_audits},
        limit=20,
    )

    assert len(hits) == 20
    assert {hit["meta"]["source_path"] for hit in hits[:3]} == {
        "F:/papers/source-0.md",
        "F:/papers/source-1.md",
        "F:/papers/source-2.md",
    }
    assert hits[2]["meta"]["matrix_field"] == "method"
